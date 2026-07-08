"""Dashboard routes — all data from real DB / stats functions."""
import io
import csv
import json
import logging
import os
from datetime import datetime, timedelta
from flask import Blueprint, render_template, redirect, url_for, send_file, request, flash, Response, session
from sqlalchemy import desc, asc
from sqlalchemy.orm import joinedload
from models import (
    SessionLocal, User, UserStatus, UserRole, SalesRole, VALID_SERVICES,
    Group, ChannelIdentity, Attempt, Grade, Question, Rubric,
    Prize, Redemption,
    QuestionType, QuestionCategory, DifficultyLevel,
    Session as SessionModel, SessionStatus,
    SpacedRepetitionQueue,
    QuestionFeedback,
    TeamReportSnapshot,
)
from services.stats import user_stats, tag_breakdown, _period_stats, period_stats_bulk
from services.session_engine import SessionEngine
from dashboard_app.auth import login_required
from dashboard_app.bot_bridge import send_telegram_message

bp = Blueprint("dashboard", __name__)
log = logging.getLogger(__name__)

_FLAG = {"mexico": "🇲🇽", "colombia": "🇨🇴", "chile": "🇨🇱", "peru": "🇵🇪"}


# ─────────────────────────────────────────────────────────────────────────────
# Public leaderboard (no login — live ranking)
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/leaderboard")
def leaderboard():
    """Public live ranking by points. Auto-refreshes every 30s."""
    db = SessionLocal()
    try:
        reps = (
            db.query(User)
            .filter(User.role.in_([UserRole.REP, UserRole.MANAGER, UserRole.ADMIN]), User.status == UserStatus.ACTIVE)
            .order_by(desc(User.points), desc(User.streak_current))
            .all()
        )
        rows = []
        for i, u in enumerate(reps, 1):
            flag = _FLAG.get(u.base_country or "", "🌎")
            rows.append({
                "rank": i,
                "name": u.name,
                "points": int(u.points or 0),
                "streak": int(u.streak_current or 0),
                "streak_best": int(u.streak_best or 0),
                "country": (u.base_country or "").capitalize(),
                "flag": flag,
            })
        return render_template("leaderboard.html", ranking=rows)
    finally:
        db.close()


@bp.route("/leaderboard/json")
def leaderboard_json():
    """JSON endpoint for live refresh (AJAX)."""
    db = SessionLocal()
    try:
        reps = (
            db.query(User)
            .filter(User.role.in_([UserRole.REP, UserRole.MANAGER, UserRole.ADMIN]), User.status == UserStatus.ACTIVE)
            .order_by(desc(User.points), desc(User.streak_current))
            .all()
        )
        rows = [
            {
                "rank": i,
                "name": u.name,
                "points": int(u.points or 0),
                "streak": int(u.streak_current or 0),
                "streak_best": int(u.streak_best or 0),
                "country": (u.base_country or "").capitalize(),
                "flag": _FLAG.get(u.base_country or "", "🌎"),
            }
            for i, u in enumerate(reps, 1)
        ]
        return Response(json.dumps({"ranking": rows}), mimetype="application/json")
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# Debug endpoint — shows raw DB state to diagnose data issues
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/debug")
@login_required
def debug():
    from models import Session as SessionModel, ChannelIdentity
    db = SessionLocal()
    try:
        users    = db.query(User).all()
        attempts = db.query(Attempt).order_by(desc(Attempt.asked_at)).limit(20).all()
        sessions = db.query(SessionModel).order_by(desc(SessionModel.date)).limit(10).all()
        identities = db.query(ChannelIdentity).all()

        rows = []
        for u in users:
            ident = db.query(ChannelIdentity).filter(ChannelIdentity.user_id == u.id).first()
            attempt_count = db.query(Attempt).filter(Attempt.user_id == u.id).count()
            grade_count   = db.query(Grade).join(Attempt, Attempt.id == Grade.attempt_id).filter(Attempt.user_id == u.id).count()
            rows.append({
                "id": u.id, "name": u.name, "status": u.status.value,
                "chat_id": ident.telegram_chat_id if ident else "NO IDENTITY",
                "attempts": attempt_count, "grades": grade_count,
            })

        recent_attempts = []
        for a in attempts:
            g = db.query(Grade).filter(Grade.attempt_id == a.id).first()
            recent_attempts.append({
                "id": a.id, "user_id": a.user_id, "question_id": a.question_id,
                "session_id": a.session_id, "is_skipped": a.is_skipped,
                "asked_at": str(a.asked_at), "answered_at": str(a.answered_at),
                "score": g.score_0_5 if g else None,
            })

        recent_sessions = [
            {"id": s.id, "user_id": s.user_id, "status": s.status.value,
             "date": str(s.date), "q_ids": s.question_ids,
             "current_idx": s.current_question_index}
            for s in sessions
        ]

        return f"""
<html><head><title>Debug</title>
<style>body{{font-family:monospace;background:#0d1117;color:#e6edf3;padding:24px;}}
h2{{color:#58a6ff;margin-top:24px;}}
table{{border-collapse:collapse;width:100%;margin-bottom:16px;}}
th,td{{border:1px solid #30363d;padding:6px 10px;text-align:left;font-size:12px;}}
th{{background:#161b22;color:#8b949e;}}
</style></head><body>
<h2>Users ({len(rows)})</h2>
<table><tr><th>ID</th><th>Name</th><th>Status</th><th>Chat ID</th><th>Attempts</th><th>Grades</th></tr>
{''.join(f"<tr><td>{r['id']}</td><td>{r['name']}</td><td>{r['status']}</td><td>{r['chat_id']}</td><td>{r['attempts']}</td><td>{r['grades']}</td></tr>" for r in rows)}
</table>
<h2>Recent Attempts ({len(recent_attempts)})</h2>
<table><tr><th>ID</th><th>User ID</th><th>Q ID</th><th>Session</th><th>Skipped</th><th>Asked At</th><th>Answered At</th><th>Score</th></tr>
{''.join(f"<tr><td>{r['id']}</td><td>{r['user_id']}</td><td>{r['question_id']}</td><td>{r['session_id']}</td><td>{r['is_skipped']}</td><td>{r['asked_at']}</td><td>{r['answered_at']}</td><td>{r['score']}</td></tr>" for r in recent_attempts)}
</table>
<h2>Recent Sessions ({len(recent_sessions)})</h2>
<table><tr><th>ID</th><th>User ID</th><th>Status</th><th>Date</th><th>Q IDs</th><th>Cur Idx</th></tr>
{''.join(f"<tr><td>{r['id']}</td><td>{r['user_id']}</td><td>{r['status']}</td><td>{r['date']}</td><td>{r['q_ids']}</td><td>{r['current_idx']}</td></tr>" for r in recent_sessions)}
</table>
<p style="color:#8b949e;font-size:11px;margin-top:24px">DB time now (utcnow): {datetime.utcnow()}</p>
</body></html>
"""
    finally:
        db.close()


def _now():
    """Naive UTC now — matches how the DB stores datetimes (datetime.utcnow())."""
    return datetime.utcnow()


def _parse_date_range(args, default_days: int = 7):
    """Parse `start`, `end`, and `preset` from query args.

    Returns (start, end, prev_start, prev_end, preset, start_str, end_str):
      - start / end: naive UTC datetimes (end is exclusive, i.e. next-day 00:00)
      - prev_start / prev_end: matching previous window of the same length
      - preset: one of '7d', '14d', '30d', 'this_week', 'last_week',
                'this_month', 'custom'
      - start_str / end_str: YYYY-MM-DD strings for <input type=date> values
    """
    now = _now()
    today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    preset = (args.get("preset") or "").strip().lower()
    start_arg = (args.get("start") or "").strip()
    end_arg = (args.get("end") or "").strip()

    def _p(s):
        try:
            return datetime.strptime(s, "%Y-%m-%d")
        except Exception:
            return None

    start = end = None
    if preset == "7d" or (not preset and not start_arg and not end_arg):
        preset = "7d"
        start = today - timedelta(days=6)
        end = today + timedelta(days=1)
    elif preset == "14d":
        start = today - timedelta(days=13)
        end = today + timedelta(days=1)
    elif preset == "30d":
        start = today - timedelta(days=29)
        end = today + timedelta(days=1)
    elif preset == "this_week":
        # Monday of current week
        start = today - timedelta(days=today.weekday())
        end = today + timedelta(days=1)
    elif preset == "last_week":
        this_mon = today - timedelta(days=today.weekday())
        start = this_mon - timedelta(days=7)
        end = this_mon
    elif preset == "this_month":
        start = today.replace(day=1)
        end = today + timedelta(days=1)
    else:
        preset = "custom"
        s = _p(start_arg) or (today - timedelta(days=default_days - 1))
        e = _p(end_arg) or today
        if e < s:
            s, e = e, s
        start = s
        end = e + timedelta(days=1)  # inclusive end

    span = end - start
    prev_end = start
    prev_start = start - span

    return (
        start,
        end,
        prev_start,
        prev_end,
        preset,
        start.strftime("%Y-%m-%d"),
        (end - timedelta(days=1)).strftime("%Y-%m-%d"),
    )


def _pct_delta(cur: float, prev: float):
    """Return (delta_abs, delta_pct or None). delta_pct None means no baseline."""
    delta = cur - prev
    if prev == 0:
        return delta, None
    return delta, round((cur - prev) / prev * 100)


def _group_filter_ids(db, group_id: int) -> list:
    """Return list of group IDs to filter by (includes descendants for country groups)."""
    g = db.query(Group).get(group_id)
    if not g:
        return []
    return g.descendant_ids(db)


def _all_users(db):
    """Return all users regardless of status (active/paused/inactive)."""
    return db.query(User).all()


# ─────────────────────────────────────────────────────────────────────────────
# Overview
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/")
@login_required
def index():
    db = SessionLocal()
    try:
        group_filter = request.args.get("group_id", "").strip()
        users = _all_users(db)
        if group_filter and group_filter.isdigit():
            gids = _group_filter_ids(db, int(group_filter))
            users = [u for u in users if u.group_id and u.group_id in gids]
        # Groups for filter: country groups first (see all sub-groups), then sub-groups
        all_groups = db.query(Group).order_by(Group.country or "", Group.name).all()
        country_first = sorted([g for g in all_groups if g.parent_group_id is None], key=lambda x: (x.country or "", x.name))
        subs = sorted([g for g in all_groups if g.parent_group_id is not None], key=lambda x: (x.country or "", x.name))
        groups = country_first + subs

        # Date range (defaults to last 7 days)
        start, end, prev_start, prev_end, preset, start_str, end_str = _parse_date_range(request.args)
        range_days = max(1, (end - start).days)

        # Team KPIs for selected window
        team = {"sent": 0, "answered": 0, "correct": 0, "expired": 0}
        team_prev = {"sent": 0, "answered": 0, "correct": 0, "expired": 0}
        for u in users:
            s = _period_stats(db, u.id, start, end)
            p = _period_stats(db, u.id, prev_start, prev_end)
            for k in team:
                team[k] += s[k]
                team_prev[k] += p[k]

        ans_pct = round(team["answered"] / team["sent"] * 100) if team["sent"] else 0
        cor_pct = round(team["correct"]  / team["answered"] * 100) if team["answered"] else 0
        ans_pct_prev = round(team_prev["answered"] / team_prev["sent"] * 100) if team_prev["sent"] else 0
        cor_pct_prev = round(team_prev["correct"]  / team_prev["answered"] * 100) if team_prev["answered"] else 0

        deltas = {
            "sent":     _pct_delta(team["sent"], team_prev["sent"]),
            "answered": _pct_delta(team["answered"], team_prev["answered"]),
            "correct":  _pct_delta(team["correct"], team_prev["correct"]),
            "expired":  _pct_delta(team["expired"], team_prev["expired"]),
            "ans_pct":  _pct_delta(ans_pct, ans_pct_prev),
            "cor_pct":  _pct_delta(cor_pct, cor_pct_prev),
        }

        # Per-user rows (selected window + prev window + all-time total)
        user_rows = []
        for u in users:
            s   = _period_stats(db, u.id, start, end)
            p   = _period_stats(db, u.id, prev_start, prev_end)
            tot = _period_stats(db, u.id, None, None)
            flag = _FLAG.get(u.base_country or "", "🌎")
            role = {"hunter": "Hunter", "farmer": "Farmer"}.get(
                u.sales_role.value if u.sales_role else "", "—"
            )
            s_cor_pct = round(s["correct"]  / s["answered"] * 100) if s["answered"] else 0
            p_cor_pct = round(p["correct"]  / p["answered"] * 100) if p["answered"] else 0
            user_rows.append({
                "id":        u.id,
                "name":      u.name,
                "flag":      flag,
                "country":   (u.base_country or "").capitalize(),
                "role":      role,
                "specs":     ", ".join(sp.capitalize() for sp in (u.specializations or [])),
                "points":    int(getattr(u, "points", 0) or 0),
                "streak":    int(getattr(u, "streak_current", 0) or 0),
                "sent":      s["sent"],
                "answered":  s["answered"],
                "correct":   s["correct"],
                "expired":   s["expired"],
                "ans_pct":   round(s["answered"] / s["sent"] * 100) if s["sent"] else 0,
                "cor_pct":   s_cor_pct,
                "prev_correct": p["correct"],
                "prev_answered": p["answered"],
                "prev_cor_pct": p_cor_pct,
                "correct_delta": s["correct"] - p["correct"],
                "cor_pct_delta": s_cor_pct - p_cor_pct,
                "total_cor": tot["correct"],
                "total_ans": tot["answered"],
            })

        # Timeseries — daily buckets when range ≤ 60 days, otherwise weekly buckets
        chart_labels, chart_answered, chart_correct = [], [], []
        if range_days <= 60:
            # Daily buckets
            cur = start
            while cur < end:
                bucket_end = cur + timedelta(days=1)
                chart_labels.append(cur.strftime("%-d %b"))
                day_ans = day_cor = 0
                for u in users:
                    s = _period_stats(db, u.id, cur, bucket_end)
                    day_ans += s["answered"]
                    day_cor += s["correct"]
                chart_answered.append(day_ans)
                chart_correct.append(day_cor)
                cur = bucket_end
        else:
            # Weekly buckets (Mon-Sun)
            first_mon = start - timedelta(days=start.weekday())
            cur = first_mon
            while cur < end:
                bucket_end = min(cur + timedelta(days=7), end)
                bucket_start = max(cur, start)
                chart_labels.append(bucket_start.strftime("%-d %b"))
                wk_ans = wk_cor = 0
                for u in users:
                    s = _period_stats(db, u.id, bucket_start, bucket_end)
                    wk_ans += s["answered"]
                    wk_cor += s["correct"]
                chart_answered.append(wk_ans)
                chart_correct.append(wk_cor)
                cur = cur + timedelta(days=7)

        # Recent activity — last 20 answered attempts within window
        user_ids = [u.id for u in users]
        recent_q = (
            db.query(Attempt, Grade, Question, User)
            .outerjoin(Grade,    Grade.attempt_id    == Attempt.id)
            .join(Question, Question.id              == Attempt.question_id)
            .join(User,     User.id                  == Attempt.user_id)
            .filter(Attempt.is_skipped == False)
            .filter(Attempt.asked_at >= start, Attempt.asked_at < end)
        )
        if user_ids:
            recent_q = recent_q.filter(Attempt.user_id.in_(user_ids))
        elif group_filter:
            recent_q = recent_q.filter(Attempt.user_id < 0)  # no reps in group → no activity
        recent = recent_q.order_by(desc(Attempt.answered_at)).limit(20).all()
        activity = []
        for attempt, grade, question, user in recent:
            activity.append({
                "user":    user.name,
                "prompt":  question.prompt[:70] + ("…" if len(question.prompt) > 70 else ""),
                "product": (question.product or "general").capitalize(),
                "score":   grade.score_0_5 if grade else None,
                "correct": bool(grade and grade.score_0_5 >= 3),
                "time":    attempt.answered_at.strftime("%-d %b %H:%M") if attempt.answered_at else "—",
            })

        # Leaderboard (top 10 by points, respects group filter)
        leaderboard = sorted(user_rows, key=lambda x: (-x["points"], -x["streak"]))[:10]

        return render_template(
            "dashboard.html",
            users=user_rows,
            team=team,
            team_prev=team_prev,
            deltas=deltas,
            ans_pct=ans_pct,
            cor_pct=cor_pct,
            ans_pct_prev=ans_pct_prev,
            cor_pct_prev=cor_pct_prev,
            chart_labels=chart_labels,
            chart_answered=chart_answered,
            chart_correct=chart_correct,
            activity=activity,
            total_reps=len(users),
            groups=groups,
            group_filter=group_filter,
            leaderboard=leaderboard,
            date_start=start_str,
            date_end=end_str,
            date_preset=preset,
            range_days=range_days,
            prev_start_label=prev_start.strftime("%-d %b"),
            prev_end_label=(prev_end - timedelta(days=1)).strftime("%-d %b"),
        )
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# Feedback from reps about questions
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/feedbacks")
@login_required
def feedbacks_page():
    db = SessionLocal()
    queue = (request.args.get("queue") or "pending").strip().lower()
    if queue not in ("pending", "completed", "all"):
        queue = "pending"
    try:
        base = (
            db.query(QuestionFeedback, User, Question)
            .join(User, User.id == QuestionFeedback.user_id)
            .outerjoin(Question, Question.id == QuestionFeedback.question_id)
        )
        pending_count = db.query(QuestionFeedback).filter(QuestionFeedback.handled.is_(False)).count()
        completed_count = db.query(QuestionFeedback).filter(QuestionFeedback.handled.is_(True)).count()

        if queue == "pending":
            q = base.filter(QuestionFeedback.handled.is_(False)).order_by(asc(QuestionFeedback.created_at))
        elif queue == "completed":
            q = base.filter(QuestionFeedback.handled.is_(True)).order_by(desc(QuestionFeedback.created_at))
        else:
            q = base.order_by(desc(QuestionFeedback.created_at))

        rows = q.limit(500).all()
        feedbacks = []
        for fb, u, qrow in rows:
            feedbacks.append(
                {
                    "id": fb.id,
                    "handled": bool(fb.handled),
                    "created_at": fb.created_at.strftime("%d/%m/%Y %H:%M") if fb.created_at else "",
                    "user_id": u.id,
                    "user_name": u.name,
                    "question_id": qrow.id if qrow else None,
                    "question_prompt": (qrow.prompt[:120] + "…")
                    if (qrow and len(qrow.prompt) > 120)
                    else (qrow.prompt if qrow else ""),
                    "comment": fb.comment,
                }
            )
        return render_template(
            "feedback.html",
            feedbacks=feedbacks,
            queue=queue,
            pending_count=pending_count,
            completed_count=completed_count,
        )
    finally:
        db.close()


@bp.route("/feedbacks/<int:fb_id>/toggle-handled", methods=["POST"])
@login_required
def feedback_toggle_handled(fb_id):
    """Mark question feedback as handled (queue done) or reopen."""
    nxt = request.form.get("next") or ""
    if not nxt.startswith("/feedbacks"):
        nxt = url_for("dashboard.feedbacks_page")
    db = SessionLocal()
    try:
        fb = db.query(QuestionFeedback).filter(QuestionFeedback.id == fb_id).first()
        if fb:
            want_handled = request.form.get("handled", "").strip().lower() in ("1", "true", "yes", "on")
            fb.handled = want_handled
            db.commit()
            flash("Feedback actualizado.", "success")
        else:
            flash("Feedback no encontrado.", "error")
    except Exception as exc:
        db.rollback()
        flash(f"Error: {exc}", "error")
    finally:
        db.close()
    return redirect(nxt)


# ─────────────────────────────────────────────────────────────────────────────
# Rep detail
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/rep/<int:user_id>")
@login_required
def rep_detail(user_id):
    db = SessionLocal()
    try:
        user = db.query(User).get(user_id)
        if not user:
            return redirect(url_for("dashboard.index"))

        # Date range (defaults to last 7 days)
        start, end, prev_start, prev_end, preset, start_str, end_str = _parse_date_range(request.args)
        range_days = max(1, (end - start).days)

        # Totals (all-time) used as context
        total = _period_stats(db, user_id, None, None)
        all_grades = (
            db.query(Grade)
            .join(Attempt, Attempt.id == Grade.attempt_id)
            .filter(Attempt.user_id == user_id)
            .all()
        )
        total_avg = (
            round(sum(g.score_0_5 for g in all_grades) / len(all_grades), 2)
            if all_grades
            else None
        )

        # Selected window
        win = _period_stats(db, user_id, start, end)
        prev = _period_stats(db, user_id, prev_start, prev_end)

        def _avg_score(s, e):
            q = (
                db.query(Grade)
                .join(Attempt, Attempt.id == Grade.attempt_id)
                .filter(Attempt.user_id == user_id, Attempt.asked_at >= s, Attempt.asked_at < e)
                .all()
            )
            return round(sum(g.score_0_5 for g in q) / len(q), 2) if q else None

        win_avg = _avg_score(start, end)
        prev_avg = _avg_score(prev_start, prev_end)

        win_ans_pct  = round(win["answered"]  / win["sent"]     * 100) if win["sent"] else 0
        win_cor_pct  = round(win["correct"]   / win["answered"] * 100) if win["answered"] else 0
        prev_ans_pct = round(prev["answered"] / prev["sent"]    * 100) if prev["sent"] else 0
        prev_cor_pct = round(prev["correct"]  / prev["answered"]* 100) if prev["answered"] else 0

        deltas = {
            "sent":     _pct_delta(win["sent"], prev["sent"]),
            "answered": _pct_delta(win["answered"], prev["answered"]),
            "correct":  _pct_delta(win["correct"], prev["correct"]),
            "expired":  _pct_delta(win["expired"], prev["expired"]),
            "ans_pct":  _pct_delta(win_ans_pct, prev_ans_pct),
            "cor_pct":  _pct_delta(win_cor_pct, prev_cor_pct),
            "avg":      _pct_delta(win_avg or 0, prev_avg or 0),
        }

        bd = tag_breakdown(user.id, db)

        # Attempts within window (capped to 60 for display)
        attempts_raw = (
            db.query(Attempt, Grade, Question)
            .outerjoin(Grade,    Grade.attempt_id == Attempt.id)
            .join(Question, Question.id           == Attempt.question_id)
            .filter(
                Attempt.user_id == user_id,
                Attempt.is_skipped == False,
                Attempt.asked_at >= start,
                Attempt.asked_at < end,
            )
            .order_by(desc(Attempt.answered_at))
            .limit(60)
            .all()
        )
        attempts = []
        for attempt, grade, question in attempts_raw:
            attempts.append({
                "prompt":  question.prompt[:80] + ("…" if len(question.prompt) > 80 else ""),
                "product": (question.product or "general").capitalize(),
                "qtype":   question.question_type.value if question.question_type else "open_ended",
                "answer":  (attempt.response_text or "")[:60] + (
                    "…" if len(attempt.response_text or "") > 60 else ""
                ),
                "score":   grade.score_0_5 if grade else None,
                "correct": bool(grade and grade.score_0_5 >= 3),
                "time":    attempt.answered_at.strftime("%-d %b %H:%M") if attempt.answered_at else "—",
            })

        # Timeseries for window: daily if ≤ 60 days, weekly otherwise
        ts_labels, ts_answered, ts_correct, ts_avg = [], [], [], []
        if range_days <= 60:
            cur = start
            while cur < end:
                nxt = cur + timedelta(days=1)
                s = _period_stats(db, user_id, cur, nxt)
                a = _avg_score(cur, nxt)
                ts_labels.append(cur.strftime("%-d %b"))
                ts_answered.append(s["answered"])
                ts_correct.append(s["correct"])
                ts_avg.append(a if a is not None else None)
                cur = nxt
        else:
            first_mon = start - timedelta(days=start.weekday())
            cur = first_mon
            while cur < end:
                bucket_end = min(cur + timedelta(days=7), end)
                bucket_start = max(cur, start)
                s = _period_stats(db, user_id, bucket_start, bucket_end)
                a = _avg_score(bucket_start, bucket_end)
                ts_labels.append(bucket_start.strftime("%-d %b"))
                ts_answered.append(s["answered"])
                ts_correct.append(s["correct"])
                ts_avg.append(a if a is not None else None)
                cur = cur + timedelta(days=7)

        # Product chart data (all-time)
        products     = sorted(bd["by_product"].items(), key=lambda x: -x[1]["answered"])
        prod_labels  = [p[0].capitalize() for p in products]
        prod_correct = [p[1]["correct"] for p in products]
        prod_wrong   = [p[1]["answered"] - p[1]["correct"] for p in products]

        flag = _FLAG.get(user.base_country or "", "🌎")
        role = {"hunter": "Hunter 🎯", "farmer": "Farmer 🌱"}.get(
            user.sales_role.value if user.sales_role else "", "—"
        )

        stats = {
            **total,
            "avg_score": total_avg,
        }

        return render_template(
            "rep_detail.html",
            rep=user,
            flag=flag,
            role=role,
            stats=stats,
            win=win,
            prev=prev,
            win_ans_pct=win_ans_pct,
            win_cor_pct=win_cor_pct,
            prev_ans_pct=prev_ans_pct,
            prev_cor_pct=prev_cor_pct,
            win_avg=win_avg,
            prev_avg=prev_avg,
            deltas=deltas,
            bd=bd,
            attempts=attempts,
            prod_labels=prod_labels,
            prod_correct=prod_correct,
            prod_wrong=prod_wrong,
            ts_labels=ts_labels,
            ts_answered=ts_answered,
            ts_correct=ts_correct,
            ts_avg=ts_avg,
            date_start=start_str,
            date_end=end_str,
            date_preset=preset,
            range_days=range_days,
            prev_start_label=prev_start.strftime("%-d %b"),
            prev_end_label=(prev_end - timedelta(days=1)).strftime("%-d %b"),
        )
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# Questions browser
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/questions")
@login_required
def questions():
    db = SessionLocal()
    try:
        all_questions = (
            db.query(Question)
            .filter(Question.active == True)
            .order_by(Question.id)
            .all()
        )

        rows = []
        for q in all_questions:
            attempts_q = (
                db.query(Attempt)
                .filter(Attempt.question_id == q.id, Attempt.is_skipped == False)
                .all()
            )
            attempts      = len(attempts_q)
            correct_count = 0
            for a in attempts_q:
                g = db.query(Grade).filter(Grade.attempt_id == a.id).first()
                if g and g.score_0_5 >= 3:
                    correct_count += 1
            pct = round(correct_count / attempts * 100) if attempts else None

            rows.append({
                "id":       q.id,
                "prompt":   q.prompt[:90] + ("…" if len(q.prompt) > 90 else ""),
                "product":  (q.product or "general").capitalize(),
                "country":  (q.country or "all"),
                "qtype":    q.question_type.value if q.question_type else "open_ended",
                "diff":     q.difficulty.value if q.difficulty else "medium",
                "attempts": attempts,
                "pct":      pct,
            })

        return render_template("questions.html", questions=rows)
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# Sync questions from data/*.json (run seed) — Railway as source of truth
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/sync-preguntas")
@login_required
def sync_questions_page():
    """Page with a button that POSTs to admin_sync_questions (avoids Method Not Allowed on GET)."""
    return render_template("sync_questions.html")


@bp.route("/admin/sync-questions", methods=["POST"])
@login_required
def admin_sync_questions():
    """Run seed_questions() to load/update all questions from data/*.json."""
    try:
        from startup_enroll import seed_questions
        seed_questions()
        flash("Preguntas sincronizadas desde data/*.json. El banco está actualizado.")
    except Exception as e:
        flash(f"Error al sincronizar: {e}", "error")
    return redirect(url_for("dashboard.questions"))


# ─────────────────────────────────────────────────────────────────────────────
# Excel export  (all data points)
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/export/excel")
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = SessionLocal()
    try:
        group_filter = request.args.get("group_id", "").strip()
        q = (
            db.query(Attempt, Grade, User, Question)
            .outerjoin(Grade,    Grade.attempt_id == Attempt.id)
            .join(User,     User.id               == Attempt.user_id)
            .join(Question, Question.id            == Attempt.question_id)
            .order_by(desc(Attempt.asked_at))
        )
        if group_filter and group_filter.isdigit():
            gids = _group_filter_ids(db, int(group_filter))
            if gids:
                q = q.filter(User.group_id.in_(gids))
        rows = q.all()

        wb = openpyxl.Workbook()

        # ── Sheet 1: All attempts ──────────────────────────────────────────
        ws = wb.active
        ws.title = "Respuestas"

        headers = [
            "Fecha", "Rep", "País", "Rol", "Especialidades",
            "Pregunta", "Producto", "País pregunta", "Tipo", "Dificultad",
            "Respuesta", "Score (0-5)", "Correcto", "Estado", "Feedback"
        ]

        # Header style
        hdr_fill   = PatternFill("solid", fgColor="4338CA")
        hdr_font   = Font(bold=True, color="FFFFFF", size=11)
        hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side  = Side(border_style="thin", color="D1D5DB")
        cell_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font      = hdr_font
            c.fill      = hdr_fill
            c.alignment = hdr_align
            c.border    = cell_border

        ws.row_dimensions[1].height = 30

        # Data rows
        green_fill = PatternFill("solid", fgColor="D1FAE5")
        red_fill   = PatternFill("solid", fgColor="FEE2E2")

        for row_idx, (attempt, grade, user, question) in enumerate(rows, 2):
            is_correct = bool(grade and grade.score_0_5 >= 3)
            row_fill   = green_fill if is_correct else red_fill

            data = [
                attempt.asked_at.strftime("%Y-%m-%d %H:%M") if attempt.asked_at else "",
                user.name,
                (user.base_country or "").capitalize(),
                user.sales_role.value.capitalize() if user.sales_role else "",
                ", ".join(s.capitalize() for s in (user.specializations or [])),
                question.prompt,
                (question.product or "").capitalize(),
                (question.country or "all"),
                question.question_type.value if question.question_type else "open_ended",
                question.difficulty.value if question.difficulty else "medium",
                attempt.response_text or "",
                grade.score_0_5 if grade else "",
                "Sí" if is_correct else "No",
                grade.pass_state.value.capitalize() if grade and grade.pass_state else "",
                (grade.feedback or "") if grade else "",
            ]

            for col, val in enumerate(data, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.border    = cell_border
                c.alignment = Alignment(vertical="top", wrap_text=(col in (6, 11, 15)))
                if col in (12, 13):
                    c.fill = row_fill

        # Column widths
        col_widths = [18, 20, 12, 12, 25, 55, 15, 14, 14, 12, 45, 12, 10, 12, 50]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # ── Sheet 2: Per-user summary ──────────────────────────────────────
        ws2 = wb.create_sheet("Resumen por rep")
        s2_headers = ["Rep", "País", "Rol", "Total enviadas", "Respondidas",
                      "% Respuesta", "Correctas", "% Acierto", "Score promedio"]
        for col, h in enumerate(s2_headers, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font      = hdr_font
            c.fill      = hdr_fill
            c.alignment = hdr_align
            c.border    = cell_border

        export_users = _all_users(db)
        if group_filter and group_filter.isdigit():
            gids = _group_filter_ids(db, int(group_filter))
            export_users = [u for u in export_users if u.group_id and u.group_id in gids]
        for row_idx, u in enumerate(export_users, 2):
            tot   = _period_stats(db, u.id, None, None)
            ugrades = (
                db.query(Grade)
                .join(Attempt, Attempt.id == Grade.attempt_id)
                .filter(Attempt.user_id == u.id)
                .all()
            )
            avg = round(sum(g.score_0_5 for g in ugrades) / len(ugrades), 2) if ugrades else ""
            ans_p = round(tot["answered"] / tot["sent"] * 100) if tot["sent"] else 0
            cor_p = round(tot["correct"]  / tot["answered"] * 100) if tot["answered"] else 0

            row_data = [
                u.name,
                (u.base_country or "").capitalize(),
                u.sales_role.value.capitalize() if u.sales_role else "",
                tot["sent"],
                tot["answered"],
                f"{ans_p}%",
                tot["correct"],
                f"{cor_p}%",
                avg,
            ]
            for col, val in enumerate(row_data, 1):
                c = ws2.cell(row=row_idx, column=col, value=val)
                c.border    = cell_border
                c.alignment = Alignment(horizontal="center")
            ws2.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")

        for i, w in enumerate([20, 12, 12, 14, 14, 12, 12, 12, 14], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = "A2"

        # ── Sheet 3: Questions performance ────────────────────────────────
        ws3 = wb.create_sheet("Preguntas")
        s3_headers = ["ID", "Pregunta", "Producto", "País", "Tipo", "Dificultad",
                      "Intentos", "Correctas", "% Acierto"]
        for col, h in enumerate(s3_headers, 1):
            c = ws3.cell(row=1, column=col, value=h)
            c.font      = hdr_font
            c.fill      = hdr_fill
            c.alignment = hdr_align
            c.border    = cell_border

        all_qs = db.query(Question).filter(Question.active == True).all()
        for row_idx, q in enumerate(all_qs, 2):
            att_q  = db.query(Attempt).filter(
                Attempt.question_id == q.id, Attempt.is_skipped == False
            ).all()
            att_n  = len(att_q)
            cor_n  = sum(
                1 for a in att_q
                if (db.query(Grade).filter(Grade.attempt_id == a.id).first() or None) and
                   db.query(Grade).filter(Grade.attempt_id == a.id).first().score_0_5 >= 3
            )
            pct = f"{round(cor_n/att_n*100)}%" if att_n else "—"

            row_data = [
                q.id, q.prompt,
                (q.product or "").capitalize(),
                (q.country or "all"),
                q.question_type.value if q.question_type else "open_ended",
                q.difficulty.value if q.difficulty else "medium",
                att_n, cor_n, pct,
            ]
            for col, val in enumerate(row_data, 1):
                c = ws3.cell(row=row_idx, column=col, value=val)
                c.border    = cell_border
                c.alignment = Alignment(vertical="top", wrap_text=(col == 2))

        for i, w in enumerate([6, 70, 15, 12, 15, 12, 10, 10, 10], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.freeze_panes = "A2"

        # Save to memory and return
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"sales_coach_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx",
        )
    finally:
        db.close()


# ── Legacy CSV kept for backward compat ───────────────────────────────────────

@bp.route("/export/csv")
@login_required
def export_csv():
    db = SessionLocal()
    try:
        rows = (
            db.query(Attempt, Grade, User, Question)
            .outerjoin(Grade,    Grade.attempt_id == Attempt.id)
            .join(User,     User.id               == Attempt.user_id)
            .join(Question, Question.id            == Attempt.question_id)
            .order_by(desc(Attempt.asked_at))
            .all()
        )
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(["Fecha", "Rep", "País", "Pregunta", "Producto", "Tipo",
                    "Respuesta", "Score", "Correcto", "Feedback"])
        for attempt, grade, user, question in rows:
            w.writerow([
                attempt.asked_at.strftime("%Y-%m-%d %H:%M") if attempt.asked_at else "",
                user.name,
                user.base_country or "",
                question.prompt[:120],
                question.product or "",
                question.question_type.value if question.question_type else "",
                (attempt.response_text or "")[:200],
                grade.score_0_5 if grade else "",
                "Sí" if (grade and grade.score_0_5 >= 3) else "No",
                (grade.feedback or "") if grade else "",
            ])
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode("utf-8")),
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"sales_coach_{datetime.utcnow().strftime('%Y%m%d')}.csv",
        )
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# CRUD — Questions
# ─────────────────────────────────────────────────────────────────────────────

# Product/service tags shown in question filters and forms. Configure via VALID_SERVICES
# env var (see models/user.py). "general" is always included as a catch-all bucket.
from models.user import VALID_SERVICES as _VALID_SERVICES
_PRODUCTS  = ["general"] + sorted(s for s in _VALID_SERVICES if s != "general")

# Countries used in rep profiles (base_country) and question targeting.
# Override via COUNTRIES env var (comma-separated), e.g. COUNTRIES="all,usa,uk,brazil"
_COUNTRIES = [c.strip() for c in os.getenv("COUNTRIES", "all,mexico,colombia,chile,peru").split(",") if c.strip()]


def _parse_form_to_question(form, question: Question = None) -> Question:
    """Build / update a Question object from a submitted form."""
    if question is None:
        question = Question()

    question.prompt        = form.get("prompt", "").strip()
    question.question_type = QuestionType(form.get("question_type", "open_ended"))
    question.product       = form.get("product", "general") or "general"
    question.country       = form.get("country", "all") or "all"
    question.difficulty    = DifficultyLevel(form.get("difficulty", "medium"))
    question.category      = QuestionCategory(form.get("category", "general"))
    question.active        = form.get("active") == "1"

    # Tags — comma-separated string → list
    raw_tags = form.get("tags", "")
    question.tags = [t.strip() for t in raw_tags.split(",") if t.strip()]

    # Correct answer (MC / Yes-No)
    question.correct_answer = form.get("correct_answer", "").strip().lower() or None

    # Choices for MC: keys A, B, C, D with text from choice_a … choice_d
    if question.question_type == QuestionType.MULTIPLE_CHOICE:
        choices = []
        for key in ["A", "B", "C", "D"]:
            text = form.get(f"choice_{key.lower()}", "").strip()
            if text:
                choices.append({"key": key, "text": text})
        question.choices = choices if choices else None
    else:
        question.choices = None

    return question


def _parse_form_to_rubric(form, question: Question) -> Rubric:
    """Build / update the Rubric for a question."""
    rubric = question.rubric or Rubric(question_id=question.id)

    rubric.ideal_answer = form.get("ideal_answer", "").strip() or None

    # must_have_concepts — one concept per line → list of plain strings
    raw = form.get("must_have_concepts", "")
    rubric.must_have_concepts = [c.strip() for c in raw.splitlines() if c.strip()]

    return rubric


@bp.route("/questions/new", methods=["GET", "POST"])
@login_required
def question_new():
    if request.method == "POST":
        db = SessionLocal()
        try:
            q = _parse_form_to_question(request.form)
            if not q.prompt:
                flash("El texto de la pregunta no puede estar vacío.", "error")
                return render_template("question_form.html",
                                       q=None, form=request.form,
                                       products=_PRODUCTS, countries=_COUNTRIES,
                                       qtypes=QuestionType, cats=QuestionCategory,
                                       diffs=DifficultyLevel, title="Nueva pregunta")
            db.add(q)
            db.flush()   # get q.id before building rubric

            rubric = _parse_form_to_rubric(request.form, q)
            rubric.question_id = q.id
            db.add(rubric)
            db.commit()
            flash("Pregunta creada correctamente.", "success")
            return redirect(url_for("dashboard.questions"))
        except Exception as exc:
            db.rollback()
            flash(f"Error al crear la pregunta: {exc}", "error")
        finally:
            db.close()

    return render_template("question_form.html",
                           q=None, form={},
                           products=_PRODUCTS, countries=_COUNTRIES,
                           qtypes=QuestionType, cats=QuestionCategory,
                           diffs=DifficultyLevel, title="Nueva pregunta")


@bp.route("/questions/<int:qid>/edit", methods=["GET", "POST"])
@login_required
def question_edit(qid):
    db = SessionLocal()
    try:
        q = db.query(Question).get(qid)
        if not q:
            flash("Pregunta no encontrada.", "error")
            return redirect(url_for("dashboard.questions"))

        if request.method == "POST":
            try:
                _parse_form_to_question(request.form, q)
                if not q.prompt:
                    flash("El texto de la pregunta no puede estar vacío.", "error")
                else:
                    rubric = _parse_form_to_rubric(request.form, q)
                    if not q.rubric:
                        rubric.question_id = q.id
                        db.add(rubric)
                    q.updated_at = datetime.utcnow()
                    db.commit()
                    flash("Pregunta actualizada.", "success")
                    return redirect(url_for("dashboard.questions"))
            except Exception as exc:
                db.rollback()
                flash(f"Error al guardar: {exc}", "error")

        # Pre-fill form values from DB
        choices_map = {}
        if q.choices:
            for c in q.choices:
                choices_map[c["key"].lower()] = c["text"]

        must_have_str = "\n".join(
            (c["concept"] if isinstance(c, dict) else c)
            for c in (q.rubric.must_have_concepts if q.rubric else [])
        )

        prefill = {
            "prompt":             q.prompt,
            "question_type":      q.question_type.value if q.question_type else "open_ended",
            "product":            q.product or "general",
            "country":            q.country or "all",
            "difficulty":         q.difficulty.value if q.difficulty else "medium",
            "category":           q.category.value if q.category else "general",
            "active":             "1" if q.active else "0",
            "tags":               ", ".join(q.tags or []),
            "correct_answer":     q.correct_answer or "",
            "choice_a":           choices_map.get("a", ""),
            "choice_b":           choices_map.get("b", ""),
            "choice_c":           choices_map.get("c", ""),
            "choice_d":           choices_map.get("d", ""),
            "ideal_answer":       (q.rubric.ideal_answer or "") if q.rubric else "",
            "must_have_concepts": must_have_str,
        }

        return render_template("question_form.html",
                               q=q, form=prefill,
                               products=_PRODUCTS, countries=_COUNTRIES,
                               qtypes=QuestionType, cats=QuestionCategory,
                               diffs=DifficultyLevel, title=f"Editar pregunta #{q.id}")
    finally:
        db.close()


@bp.route("/questions/<int:qid>/delete", methods=["POST"])
@login_required
def question_delete(qid):
    db = SessionLocal()
    try:
        q = db.query(Question).get(qid)
        if q:
            db.delete(q)
            db.commit()
            flash(f"Pregunta #{qid} eliminada.", "success")
        else:
            flash("Pregunta no encontrada.", "error")
    except Exception as exc:
        db.rollback()
        flash(f"Error al eliminar: {exc}", "error")
    finally:
        db.close()
    return redirect(url_for("dashboard.questions"))


@bp.route("/questions/<int:qid>/toggle", methods=["POST"])
@login_required
def question_toggle(qid):
    db = SessionLocal()
    try:
        q = db.query(Question).get(qid)
        if q:
            q.active = not q.active
            db.commit()
            state = "activada" if q.active else "desactivada"
            flash(f"Pregunta #{qid} {state}.", "success")
    except Exception as exc:
        db.rollback()
        flash(f"Error: {exc}", "error")
    finally:
        db.close()
    return redirect(url_for("dashboard.questions"))


# ═════════════════════════════════════════════════════════════════════════════
# USERS MODULE
# ═════════════════════════════════════════════════════════════════════════════

_SPECIALIZATIONS = sorted(VALID_SERVICES)
def _sync_enrolled_json(db) -> None:
    """Rewrite enrolled_users.json to match the current DB state."""
    from startup_enroll import sync_enrolled_json
    sync_enrolled_json(db)


# ─────────────────────────────────────────────────────────────────────────────
# Groups CRUD
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/groups")
@login_required
def groups():
    db = SessionLocal()
    try:
        all_groups = db.query(Group).order_by(Group.name).all()
        # Country groups first, then sub-groups
        all_groups = sorted([g for g in all_groups if g.parent_group_id is None], key=lambda x: (x.country or "", x.name)) + sorted([g for g in all_groups if g.parent_group_id is not None], key=lambda x: (x.country or "", x.name))
        managers = db.query(User).filter(
            User.role.in_([UserRole.MANAGER, UserRole.ADMIN])
        ).order_by(User.name).all()
        rows = []
        for g in all_groups:
            gids = g.descendant_ids(db)
            user_count = db.query(User).filter(User.group_id.in_(gids)).count() if gids else 0
            manager = db.query(User).get(g.manager_id) if g.manager_id else None
            country = (g.country or "").capitalize() if g.country else "—"
            parent = db.query(Group).get(g.parent_group_id) if g.parent_group_id else None
            rows.append({
                "id": g.id,
                "name": g.name,
                "country": country,
                "parent": parent.name if parent else None,
                "manager": manager.name if manager else None,
                "user_count": user_count,
            })
        return render_template("groups.html", groups=rows, managers=managers)
    finally:
        db.close()


@bp.route("/groups/new", methods=["GET", "POST"])
@login_required
def group_new():
    if request.method == "POST":
        db = SessionLocal()
        try:
            name = request.form.get("name", "").strip()
            if not name:
                flash("El nombre del grupo es obligatorio.", "error")
                managers = db.query(User).filter(User.role.in_([UserRole.MANAGER, UserRole.ADMIN])).order_by(User.name).all()
                parent_groups = db.query(Group).filter(Group.parent_group_id.is_(None)).order_by(Group.name).all()
                return render_template("group_form.html", g=None, form=request.form, managers=managers, parent_groups=parent_groups, title="Nuevo grupo")
            mid = request.form.get("manager_id", "").strip()
            pid = request.form.get("parent_group_id", "").strip()
            country = request.form.get("country", "").strip().lower() or None
            g = Group(
                name=name,
                country=country,
                parent_group_id=int(pid) if pid and pid.isdigit() else None,
                manager_id=int(mid) if mid and mid.isdigit() else None,
            )
            db.add(g)
            db.commit()
            flash(f"Grupo '{name}' creado.", "success")
            return redirect(url_for("dashboard.groups"))
        except Exception as exc:
            db.rollback()
            flash(f"Error al crear grupo: {exc}", "error")
        finally:
            db.close()

    db = SessionLocal()
    try:
        managers = db.query(User).filter(User.role.in_([UserRole.MANAGER, UserRole.ADMIN])).order_by(User.name).all()
        parent_groups = db.query(Group).filter(Group.parent_group_id.is_(None)).order_by(Group.name).all()
        return render_template("group_form.html", g=None, form={}, managers=managers, parent_groups=parent_groups, title="Nuevo grupo")
    finally:
        db.close()


@bp.route("/groups/<int:gid>/edit", methods=["GET", "POST"])
@login_required
def group_edit(gid):
    db = SessionLocal()
    try:
        g = db.query(Group).get(gid)
        if not g:
            flash("Grupo no encontrado.", "error")
            return redirect(url_for("dashboard.groups"))

        if request.method == "POST":
            name = request.form.get("name", "").strip()
            if not name:
                flash("El nombre del grupo es obligatorio.", "error")
                managers = db.query(User).filter(User.role.in_([UserRole.MANAGER, UserRole.ADMIN])).order_by(User.name).all()
                parent_groups = db.query(Group).filter(Group.parent_group_id.is_(None)).filter(Group.id != gid).order_by(Group.name).all()
                return render_template("group_form.html", g=g, form=request.form, managers=managers, parent_groups=parent_groups, title=f"Editar — {g.name}")
            mid = request.form.get("manager_id", "").strip()
            pid = request.form.get("parent_group_id", "").strip()
            country = request.form.get("country", "").strip().lower() or None
            g.name = name
            g.country = country
            g.parent_group_id = int(pid) if pid and pid.isdigit() else None
            g.manager_id = int(mid) if mid and mid.isdigit() else None
            db.commit()
            flash(f"Grupo '{name}' actualizado.", "success")
            return redirect(url_for("dashboard.groups"))

        managers = db.query(User).filter(User.role.in_([UserRole.MANAGER, UserRole.ADMIN])).order_by(User.name).all()
        parent_groups = db.query(Group).filter(Group.parent_group_id.is_(None)).filter(Group.id != gid).order_by(Group.name).all()
        prefill = {
            "name": g.name,
            "country": g.country or "",
            "parent_group_id": str(g.parent_group_id) if g.parent_group_id else "",
            "manager_id": str(g.manager_id) if g.manager_id else "",
        }
        return render_template("group_form.html", g=g, form=prefill, managers=managers, parent_groups=parent_groups, title=f"Editar — {g.name}")
    finally:
        db.close()


@bp.route("/groups/<int:gid>/delete", methods=["POST"])
@login_required
def group_delete(gid):
    db = SessionLocal()
    try:
        g = db.query(Group).get(gid)
        if g:
            # Unassign users from this group
            db.query(User).filter(User.group_id == gid).update({User.group_id: None})
            db.delete(g)
            db.commit()
            flash("Grupo eliminado.", "success")
        else:
            flash("Grupo no encontrado.", "error")
    except Exception as exc:
        db.rollback()
        flash(f"Error al eliminar: {exc}", "error")
    finally:
        db.close()
    return redirect(url_for("dashboard.groups"))


# ─────────────────────────────────────────────────────────────────────────────
# Prizes CRUD
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/prizes")
@login_required
def prizes():
    db = SessionLocal()
    try:
        all_prizes = db.query(Prize).order_by(Prize.points_cost).all()
        rows = [
            {
                "id": p.id,
                "name": p.name,
                "description": p.description,
                "points_cost": p.points_cost,
                "quantity_available": p.quantity_available,
                "active": p.active,
            }
            for p in all_prizes
        ]
        return render_template("prizes.html", prizes=rows)
    finally:
        db.close()


@bp.route("/prizes/new", methods=["GET", "POST"])
@login_required
def prize_new():
    if request.method == "POST":
        db = SessionLocal()
        try:
            name = request.form.get("name", "").strip()
            if not name:
                flash("El nombre es obligatorio.", "error")
                return render_template("prize_form.html", p=None, form=request.form, title="Nuevo premio")
            pts = request.form.get("points_cost", "").strip()
            qty = request.form.get("quantity_available", "").strip()
            p = Prize(
                name=name,
                description=request.form.get("description", "").strip() or None,
                points_cost=int(pts) if pts and pts.isdigit() else 0,
                quantity_available=int(qty) if qty and qty.isdigit() else None,
                active=bool(request.form.get("active")),
            )
            if p.points_cost < 1:
                flash("Los puntos deben ser al menos 1.", "error")
                return render_template("prize_form.html", p=None, form=request.form, title="Nuevo premio")
            db.add(p)
            db.commit()
            flash(f"Premio '{name}' creado.", "success")
            return redirect(url_for("dashboard.prizes"))
        except Exception as exc:
            db.rollback()
            flash(f"Error: {exc}", "error")
        finally:
            db.close()
    return render_template("prize_form.html", p=None, form={}, title="Nuevo premio")


@bp.route("/prizes/<int:pid>/edit", methods=["GET", "POST"])
@login_required
def prize_edit(pid):
    db = SessionLocal()
    try:
        p = db.query(Prize).get(pid)
        if not p:
            flash("Premio no encontrado.", "error")
            return redirect(url_for("dashboard.prizes"))
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            if not name:
                flash("El nombre es obligatorio.", "error")
                return render_template("prize_form.html", p=p, form=request.form, title=f"Editar — {p.name}")
            pts = request.form.get("points_cost", "").strip()
            qty = request.form.get("quantity_available", "").strip()
            p.name = name
            p.description = request.form.get("description", "").strip() or None
            p.points_cost = int(pts) if pts and pts.isdigit() else p.points_cost
            p.quantity_available = int(qty) if qty and qty.isdigit() else None
            if qty == "":
                p.quantity_available = None
            p.active = bool(request.form.get("active"))
            db.commit()
            flash(f"Premio '{name}' actualizado.", "success")
            return redirect(url_for("dashboard.prizes"))
        prefill = {
            "name": p.name,
            "description": p.description or "",
            "points_cost": p.points_cost,
            "quantity_available": p.quantity_available if p.quantity_available is not None else "",
            "active": "1" if p.active else "0",
        }
        return render_template("prize_form.html", p=p, form=prefill, title=f"Editar — {p.name}")
    finally:
        db.close()


@bp.route("/prizes/<int:pid>/delete", methods=["POST"])
@login_required
def prize_delete(pid):
    db = SessionLocal()
    try:
        p = db.query(Prize).get(pid)
        if p:
            db.delete(p)
            db.commit()
            flash("Premio eliminado.", "success")
        else:
            flash("Premio no encontrado.", "error")
    except Exception as exc:
        db.rollback()
        flash(f"Error: {exc}", "error")
    finally:
        db.close()
    return redirect(url_for("dashboard.prizes"))


# ─────────────────────────────────────────────────────────────────────────────
# Redemption page
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/redeem", methods=["GET", "POST"])
def redeem():
    from services.redemption import redeem_prize

    db = SessionLocal()
    try:
        token = request.args.get("token", "").strip()
        user_id = None
        user_points = 0

        # Rep view: token in URL (no login required)
        if token:
            user = db.query(User).filter(User.redeem_token == token).first()
            if not user:
                flash("Enlace inválido o expirado.", "error")
                return redirect(url_for("login"))
            user_id = user.id
            user_points = int(user.points or 0)

            # Rep POST: self-redemption
            if request.method == "POST":
                pid = request.form.get("prize_id", "").strip()
                if pid and pid.isdigit():
                    ok, msg = redeem_prize(db, user_id, int(pid))
                    flash(msg, "success" if ok else "error")
                return redirect(url_for("dashboard.redeem", token=token))

            # Rep GET: show prizes and redeem form
            prizes = db.query(Prize).filter(Prize.active == True).all()
            prize_rows = [
                {
                    "id": p.id, "name": p.name, "description": p.description,
                    "points_cost": p.points_cost, "quantity_available": p.quantity_available,
                    "available": p.quantity_available is None or p.quantity_available > 0,
                }
                for p in prizes
            ]
            return render_template(
                "redeem.html",
                user_points=user_points,
                user_id=user_id,
                prizes=prize_rows,
                reps=None,
                redemptions=None,
            )

        # Admin view: logged in, no token
        if session.get("authenticated"):
            if request.method == "POST":
                uid = request.form.get("user_id", "").strip()
                pid = request.form.get("prize_id", "").strip()
                if uid and pid and uid.isdigit() and pid.isdigit():
                    ok, msg = redeem_prize(db, int(uid), int(pid))
                    flash(msg, "success" if ok else "error")
                return redirect(url_for("dashboard.redeem"))

            reps = db.query(User).filter(User.role == UserRole.REP, User.status == UserStatus.ACTIVE).order_by(User.name).all()
            reps_data = [{"id": r.id, "name": r.name, "points": int(r.points or 0)} for r in reps]
            prizes = db.query(Prize).filter(Prize.active == True).all()
            prize_rows = [
                {
                    "id": p.id, "name": p.name, "description": p.description,
                    "points_cost": p.points_cost, "quantity_available": p.quantity_available,
                    "available": p.quantity_available is None or p.quantity_available > 0,
                }
                for p in prizes
            ]
            redemptions = (
                db.query(Redemption, User, Prize)
                .join(User, Redemption.user_id == User.id)
                .join(Prize, Redemption.prize_id == Prize.id)
                .order_by(desc(Redemption.redeemed_at))
                .limit(50)
                .all()
            )
            hist = [
                {"user_name": u.name, "prize_name": pr.name, "points_spent": r.points_spent, "redeemed_at": r.redeemed_at.strftime("%Y-%m-%d %H:%M") if r.redeemed_at else ""}
                for r, u, pr in redemptions
            ]
            return render_template(
                "redeem.html",
                user_points=0,
                user_id=None,
                prizes=prize_rows,
                reps=reps_data,
                redemptions=hist,
            )

        # No token, not logged in
        flash("Inicia sesión o usa tu enlace de canje.", "error")
        return redirect(url_for("login"))
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────────────────────
# Invite links (one-time enrollment)
# ─────────────────────────────────────────────────────────────────────────────

def _invite_link(code: str = None, use_api: bool = True) -> str:
    from config import Config
    code = code or Config.ENROLL_CODE
    username = (Config.TELEGRAM_BOT_USERNAME or "").strip()
    if not username and use_api:
        from dashboard_app.bot_bridge import get_bot_username
        username = get_bot_username()
    if username:
        return f"https://t.me/{username}?start={code}"
    return f"https://t.me/Coach_growth99_bot?start={code}"


@bp.route("/invites")
@login_required
def invites():
    from config import Config
    link = _invite_link()
    return render_template("invites.html", link=link, base_url=Config.BASE_URL.rstrip("/"))


@bp.route("/enroll", methods=["GET", "POST"])
def enroll_redirect():
    """
    Public self-enrollment: user fills form (name, email, Telegram ID), we create User (PENDING) + ChannelIdentity.
    Then they open the bot and send /start to link chat and start (after admin approval).
    """
    from config import Config
    # Don't call Telegram API on every page load (avoids timeout/crash on Railway)
    link = _invite_link(use_api=False)

    if request.args.get("redirect"):
        return redirect(link)

    def _get_groups_safe():
        """Return groups list or [] if DB fails (e.g. Railway volume not mounted)."""
        try:
            db = SessionLocal()
            try:
                return db.query(Group).order_by(Group.name).all()
            finally:
                db.close()
        except Exception as e:
            log.warning("enroll: could not load groups: %s", e)
            return []

    try:
        if request.method == "POST":
            db = SessionLocal()
            try:
                name = (request.form.get("name") or "").strip()
                email = (request.form.get("email") or "").strip()
                raw_tg = (request.form.get("telegram_id") or "").strip().replace(" ", "")
                group_id = request.form.get("group_id", "").strip()
                role_key = (request.form.get("role") or "farmer").strip().lower()

                errors = []
                if not name:
                    errors.append("El nombre es obligatorio.")
                if not email:
                    errors.append("El correo es obligatorio.")
                else:
                    # Optionally restrict enrollment to your corporate domain by setting
                    # ENROLL_EMAIL_DOMAIN in .env (e.g. "yourcompany.com").
                    _domain = os.getenv("ENROLL_EMAIL_DOMAIN", "").strip().lower().lstrip("@")
                    if _domain and not email.lower().endswith(f"@{_domain}"):
                        errors.append(f"El correo debe ser una dirección @{_domain}.")
                if not raw_tg:
                    errors.append("El ID de Telegram es obligatorio.")
                elif not raw_tg.isdigit():
                    errors.append("El ID de Telegram debe ser un número (obtén el tuyo escribiendo /start al bot).")

                if errors:
                    groups = _get_groups_safe()
                    return render_template(
                        "enroll.html",
                        link=link,
                        groups=groups,
                        errors=errors,
                        form=request.form,
                    )

                telegram_user_id = str(int(raw_tg))
                existing = db.query(ChannelIdentity).filter(
                    ChannelIdentity.telegram_user_id == telegram_user_id
                ).first()
                if existing:
                    groups = _get_groups_safe()
                    return render_template(
                        "enroll.html",
                        link=link,
                        groups=groups,
                        errors=["Ese ID de Telegram ya está inscrito. Si es tuyo, abre el bot y escribe /start."],
                        form=request.form,
                    )

                group_id = int(group_id) if group_id and group_id.isdigit() else None
                role_map = {"farmer": UserRole.REP, "hunter": UserRole.REP, "manager": UserRole.MANAGER}
                sales_map = {"farmer": SalesRole.FARMER, "hunter": SalesRole.HUNTER, "manager": None}
                user_role = role_map.get(role_key, UserRole.REP)
                sales_role = sales_map.get(role_key, SalesRole.FARMER)

                manager_id = None
                if group_id:
                    grp = db.query(Group).get(group_id)
                    if grp:
                        manager_id = grp.manager_id

                u = User(
                    name=name,
                    email=email,
                    role=user_role,
                    status=UserStatus.PENDING,
                    group_id=group_id,
                    manager_id=manager_id,
                    sales_role=sales_role,
                )
                db.add(u)
                db.flush()
                ident = ChannelIdentity(
                    user_id=u.id,
                    channel="telegram",
                    telegram_user_id=telegram_user_id,
                    telegram_username=None,
                    telegram_chat_id=None,
                )
                db.add(ident)
                db.commit()
                _sync_enrolled_json(db)
                return render_template("enroll.html", link=link, groups=[], success=True, enroll_name=name)
            except Exception as exc:
                db.rollback()
                log.exception("enroll failed: %s", exc)
                groups = _get_groups_safe()
                return render_template(
                    "enroll.html",
                    link=link,
                    groups=groups,
                    errors=["Error al registrar. Intenta de nuevo o contacta a tu manager."],
                    form=request.form,
                )
            finally:
                db.close()

        groups = _get_groups_safe()
        return render_template("enroll.html", link=link, groups=groups, form=None)
    except Exception as exc:
        log.exception("enroll page failed: %s", exc)
        return (
            "<!DOCTYPE html><html><head><meta charset='utf-8'/><title>Error</title></head><body style='font-family:sans-serif;max-width:480px;margin:2rem auto;padding:1rem;'>"
            "<h1>Error al cargar la página</h1><p>Intenta de nuevo en unos segundos. Si sigue fallando, contacta a tu manager.</p>"
            "</body></html>",
            503,
            {"Content-Type": "text/html; charset=utf-8"},
        )



# ─────────────────────────────────────────────────────────────────────────────
# Users
# ─────────────────────────────────────────────────────────────────────────────

@bp.route("/users")
@login_required
def users():
    db = SessionLocal()
    try:
        group_filter = request.args.get("group_id", "").strip()
        status_filter = request.args.get("status", "").strip()
        query = (
            db.query(User)
            .options(
                joinedload(User.group).joinedload(Group.manager),
            )
            .order_by(User.name)
        )
        if group_filter and group_filter.isdigit():
            gids = _group_filter_ids(db, int(group_filter))
            query = query.filter(User.group_id.in_(gids)) if gids else query.filter(User.group_id < 0)
        if status_filter == "pending":
            query = query.filter(User.status == UserStatus.PENDING)
        all_users = query.all()

        all_grps = db.query(Group).order_by(Group.name).all()
        groups = sorted([g for g in all_grps if g.parent_group_id is None], key=lambda x: (x.country or "", x.name)) + \
                 sorted([g for g in all_grps if g.parent_group_id is not None], key=lambda x: (x.country or "", x.name))

        user_ids = [u.id for u in all_users]
        identities = (
            db.query(ChannelIdentity).filter(ChannelIdentity.user_id.in_(user_ids)).all()
            if user_ids else []
        )
        ident_by_user = {i.user_id: i for i in identities}
        totals_by_user = period_stats_bulk(db, user_ids, None, None) if user_ids else {}

        rows = []
        for u in all_users:
            ident = ident_by_user.get(u.id)
            group = u.group
            manager = (group.manager if group else None) or (db.query(User).get(u.manager_id) if u.manager_id else None)
            tot = totals_by_user.get(u.id, {"sent": 0, "correct": 0})
            rows.append({
                "id":          u.id,
                "name":        u.name,
                "email":       u.email or "—",
                "chat_id":     ident.telegram_chat_id  if ident else "—",
                "username":    ident.telegram_username if ident else "—",
                "role":        u.role.value,
                "status":      u.status.value,
                "sales_role":  u.sales_role.value if u.sales_role else "—",
                "group":       group.name if group else "—",
                "group_id":    u.group_id,
                "manager":     manager.name if manager else "—",
                "country":     (u.base_country or "").capitalize() or "—",
                "specs":       ", ".join(s.capitalize() for s in (u.specializations or [])) or "—",
                "total_sent":  tot["sent"],
                "total_cor":   tot["correct"],
            })
        return render_template("users.html", users=rows, groups=groups, group_filter=group_filter, status_filter=status_filter)
    except Exception as exc:
        log.exception("users list failed: %s", exc)
        flash("Error al cargar la lista de usuarios. Revisa los logs.", "error")
        return render_template("users.html", users=[], groups=[], group_filter="", status_filter=request.args.get("status", ""))
    finally:
        db.close()


def _send_welcome_and_first_question(db, chat_id: str, user_id: int, user_name: str):
    """Send welcome + journey message and first training question to a newly added user.
    Returns (True, None) on success, (False, error_description) on failure.
    """
    app_name = os.getenv("APP_NAME", "Sales Coach")
    welcome = (
        f"👋 *¡Bienvenido/a a {app_name}, {user_name}!*\n\n"
        "Soy tu coach de ventas. Así funciona tu entrenamiento:\n\n"
        "📅 *Entrenamiento diario*\n"
        "• Recibirás preguntas en horario laboral (L–V).\n"
        "• También puedes escribir /empezar cuando quieras para recibir tu siguiente pregunta.\n\n"
        "💬 *Responde* con texto o con una nota de voz 🎙️; obtendrás retroalimentación al instante.\n\n"
        "⭐ *Puntos y rachas*\n"
        "• Sumas puntos por cada respuesta correcta y mantienes rachas diarias.\n\n"
        "Aquí va tu primera pregunta. Responde cuando estés listo/a."
    )
    ok, err = send_telegram_message(chat_id, welcome, parse_mode="Markdown")
    if not ok:
        return False, err

    engine = SessionEngine(db)
    session = engine.get_active_session(user_id)
    if not session:
        session = engine.create_daily_session(user_id)
    if session.status == SessionStatus.PENDING:
        engine.start_session(session.id)
        db.refresh(session)

    from handlers.conversations import _format_question
    question = engine.get_current_question(session)
    if question:
        q_text = _format_question(question, 1, len(session.question_ids or []))
        ok2, err2 = send_telegram_message(chat_id, q_text, parse_mode="Markdown")
        if not ok2:
            return False, err2
    return True, None


@bp.route("/users/new", methods=["GET", "POST"])
@login_required
def user_new():
    if request.method == "POST":
        db = SessionLocal()
        try:
            f = request.form
            # Create User
            u = User(
                name   = f.get("name", "").strip(),
                email  = f.get("email", "").strip() or None,
                role   = UserRole(f.get("role", "rep")),
                status = UserStatus.ACTIVE if f.get("active") == "1" else UserStatus.INACTIVE,
            )
            raw_sr = f.get("sales_role", "")
            if raw_sr and raw_sr in SalesRole.__members__:
                u.sales_role = SalesRole[raw_sr]
            u.base_country    = f.get("base_country", "").lower() or None
            u.specializations = [s for s in f.getlist("specializations") if s]

            if not u.name:
                flash("El nombre es obligatorio.", "error")
                all_grps = db.query(Group).order_by(Group.name).all()
                groups = sorted([g for g in all_grps if g.parent_group_id is None], key=lambda x: (x.country or "", x.name)) + sorted([g for g in all_grps if g.parent_group_id is not None], key=lambda x: (x.country or "", x.name))
                return render_template("user_form.html", u=None, form=f,
                                       specs=_SPECIALIZATIONS, groups=groups, title="Nuevo usuario")

            gid = f.get("group_id", "").strip()
            u.group_id = int(gid) if gid and gid.isdigit() else None
            db.add(u)
            db.flush()

            # Create ChannelIdentity
            chat_id = f.get("chat_id", "").strip()
            if chat_id:
                ident = ChannelIdentity(
                    user_id           = u.id,
                    telegram_chat_id  = chat_id,
                    telegram_user_id  = f.get("telegram_user_id", "").strip() or chat_id,
                    telegram_username = f.get("username", "").strip() or None,
                )
                db.add(ident)

            db.commit()
            _sync_enrolled_json(db)

            # Send welcome + first question from the bot (only if ACTIVE and has chat_id)
            if chat_id and u.status == UserStatus.ACTIVE:
                ok, send_err = _send_welcome_and_first_question(db, chat_id, u.id, u.name)
                if not ok:
                    flash(
                        f"Usuario '{u.name}' creado, pero no se pudo enviar el mensaje de bienvenida al bot: {send_err}. "
                        "El usuario debe abrir el bot en Telegram y escribir /start o /empezar primero.",
                        "warning",
                    )
                else:
                    flash(f"Usuario '{u.name}' creado. Se envió mensaje de bienvenida y primera pregunta al bot.", "success")
            else:
                flash(f"Usuario '{u.name}' creado correctamente.", "success")
            return redirect(url_for("dashboard.users"))
        except Exception as exc:
            db.rollback()
            flash(f"Error al crear usuario: {exc}", "error")
        finally:
            db.close()

    db = SessionLocal()
    try:
        all_grps = db.query(Group).order_by(Group.name).all()
        groups = sorted([g for g in all_grps if g.parent_group_id is None], key=lambda x: (x.country or "", x.name)) + sorted([g for g in all_grps if g.parent_group_id is not None], key=lambda x: (x.country or "", x.name))
        return render_template("user_form.html", u=None, form={},
                               specs=_SPECIALIZATIONS, groups=groups, title="Nuevo usuario")
    finally:
        db.close()


@bp.route("/users/<int:uid>/edit", methods=["GET", "POST"])
@login_required
def user_edit(uid):
    db = SessionLocal()
    try:
        u = db.query(User).get(uid)
        if not u:
            flash("Usuario no encontrado.", "error")
            return redirect(url_for("dashboard.users"))

        ident = db.query(ChannelIdentity).filter(ChannelIdentity.user_id == uid).first()

        if request.method == "POST":
            f = request.form
            try:
                u.name   = f.get("name", "").strip() or u.name
                u.email  = f.get("email", "").strip() or None
                u.role   = UserRole(f.get("role", "rep"))
                u.status = UserStatus.ACTIVE if f.get("active") == "1" else UserStatus.INACTIVE

                raw_sr = f.get("sales_role", "").upper()
                u.sales_role = SalesRole[raw_sr] if raw_sr in SalesRole.__members__ else None
                u.base_country    = f.get("base_country", "").lower() or None
                u.specializations = [s for s in f.getlist("specializations") if s]
                gid = f.get("group_id", "").strip()
                u.group_id = int(gid) if gid and gid.isdigit() else None

                chat_id = f.get("chat_id", "").strip()
                if ident:
                    ident.telegram_chat_id  = chat_id or ident.telegram_chat_id
                    ident.telegram_user_id  = f.get("telegram_user_id", "").strip() or ident.telegram_user_id
                    ident.telegram_username = f.get("username", "").strip() or None
                elif chat_id:
                    db.add(ChannelIdentity(
                        user_id           = u.id,
                        telegram_chat_id  = chat_id,
                        telegram_user_id  = f.get("telegram_user_id", "").strip() or chat_id,
                        telegram_username = f.get("username", "").strip() or None,
                    ))

                db.commit()
                _sync_enrolled_json(db)
                flash(f"Usuario '{u.name}' actualizado.", "success")
                return redirect(url_for("dashboard.users"))
            except Exception as exc:
                db.rollback()
                flash(f"Error al guardar: {exc}", "error")

        prefill = {
            "name":            u.name,
            "email":           u.email or "",
            "role":            u.role.value,
            "active":          "1" if u.status == UserStatus.ACTIVE else "0",
            "sales_role":      u.sales_role.name if u.sales_role else "",
            "base_country":    u.base_country or "",
            "specializations": u.specializations or [],
            "group_id":        str(u.group_id) if u.group_id else "",
            "chat_id":         ident.telegram_chat_id  if ident else "",
            "telegram_user_id": ident.telegram_user_id if ident else "",
            "username":        ident.telegram_username if ident else "",
        }
        all_grps = db.query(Group).order_by(Group.name).all()
        groups = sorted([g for g in all_grps if g.parent_group_id is None], key=lambda x: (x.country or "", x.name)) + sorted([g for g in all_grps if g.parent_group_id is not None], key=lambda x: (x.country or "", x.name))
        return render_template("user_form.html", u=u, form=prefill,
                               specs=_SPECIALIZATIONS, groups=groups, title=f"Editar — {u.name}")
    finally:
        db.close()


@bp.route("/users/<int:uid>/delete", methods=["POST"])
@login_required
def user_delete(uid):
    db = SessionLocal()
    try:
        u = db.query(User).get(uid)
        if u:
            # Remove spaced_repetition_queue rows that reference this user's attempts (FK would block delete)
            db.query(SpacedRepetitionQueue).filter(SpacedRepetitionQueue.user_id == uid).delete(synchronize_session=False)
            db.delete(u)
            db.commit()
            _sync_enrolled_json(db)
            flash(f"Usuario eliminado.", "success")
        else:
            flash("Usuario no encontrado.", "error")
    except Exception as exc:
        db.rollback()
        flash(f"Error al eliminar: {exc}", "error")
    finally:
        db.close()
    return redirect(url_for("dashboard.users"))


@bp.route("/users/<int:uid>/toggle", methods=["POST"])
@login_required
def user_toggle(uid):
    db = SessionLocal()
    try:
        u = db.query(User).get(uid)
        if u:
            u.status = UserStatus.INACTIVE if u.status == UserStatus.ACTIVE else UserStatus.ACTIVE
            db.commit()
            _sync_enrolled_json(db)
            state = "activado" if u.status == UserStatus.ACTIVE else "desactivado"
            flash(f"Usuario {state}.", "success")
    except Exception as exc:
        db.rollback()
        flash(f"Error: {exc}", "error")
    finally:
        db.close()
    return redirect(url_for("dashboard.users"))


@bp.route("/users/<int:uid>/approve", methods=["POST"])
@login_required
def user_approve(uid):
    db = SessionLocal()
    try:
        u = db.query(User).get(uid)
        if u and u.status == UserStatus.PENDING:
            u.status = UserStatus.ACTIVE
            db.commit()
            _sync_enrolled_json(db)
            ident = db.query(ChannelIdentity).filter(ChannelIdentity.user_id == uid).first()
            if ident and ident.telegram_chat_id:
                from dashboard_app.bot_bridge import send_telegram_message
                ok, _ = send_telegram_message(
                    ident.telegram_chat_id,
                    f"✅ ¡Tu usuario fue aprobado! Ya puedes usar /start o /empezar para comenzar tu entrenamiento en {os.getenv('APP_NAME', 'Sales Coach')}."
                )
            flash(f"Usuario '{u.name}' aprobado.", "success")
        else:
            flash("Usuario no encontrado o no está pendiente.", "error")
    except Exception as exc:
        db.rollback()
        flash(f"Error: {exc}", "error")
    finally:
        db.close()
    return redirect(request.referrer or url_for("dashboard.users"))


@bp.route("/users/<int:uid>/reject", methods=["POST"])
@login_required
def user_reject(uid):
    db = SessionLocal()
    try:
        u = db.query(User).get(uid)
        if u and u.status == UserStatus.PENDING:
            u.status = UserStatus.INACTIVE
            db.commit()
            _sync_enrolled_json(db)
            ident = db.query(ChannelIdentity).filter(ChannelIdentity.user_id == uid).first()
            if ident and ident.telegram_chat_id:
                from dashboard_app.bot_bridge import send_telegram_message
                ok, _ = send_telegram_message(
                    ident.telegram_chat_id,
                    "Tu solicitud de inscripción fue rechazada. Contacta a tu manager si necesitas ayuda."
                )
            flash(f"Usuario '{u.name}' rechazado.", "success")
        else:
            flash("Usuario no encontrado o no está pendiente.", "error")
    except Exception as exc:
        db.rollback()
        flash(f"Error: {exc}", "error")
    finally:
        db.close()
    return redirect(request.referrer or url_for("dashboard.users"))


# ═════════════════════════════════════════════════════════════════════════════
# SEND QUESTIONS MODULE
# ═════════════════════════════════════════════════════════════════════════════

@bp.route("/send", methods=["GET", "POST"])
@login_required
def send_questions():
    from dashboard_app.bot_bridge import send_telegram_message, arm_bot_state

    db = SessionLocal()
    try:
        # Fetch users with a telegram_chat_id
        identities = (
            db.query(ChannelIdentity, User)
            .join(User, User.id == ChannelIdentity.user_id)
            .filter(ChannelIdentity.telegram_chat_id.isnot(None))
            .order_by(User.name)
            .all()
        )
        user_list = [
            {
                "id":      u.id,
                "name":    u.name,
                "chat_id": ident.telegram_chat_id,
                "country": (u.base_country or "").capitalize(),
                "role":    u.sales_role.value if u.sales_role else "",
            }
            for ident, u in identities
        ]

        # All active questions for selection
        questions_all = (
            db.query(Question)
            .filter(Question.active == True)
            .order_by(Question.product, Question.id)
            .all()
        )
        q_list = [
            {
                "id":      q.id,
                "prompt":  q.prompt[:100] + ("…" if len(q.prompt) > 100 else ""),
                "product": (q.product or "general").capitalize(),
                "qtype":   q.question_type.value if q.question_type else "open_ended",
                "diff":    q.difficulty.value if q.difficulty else "medium",
            }
            for q in questions_all
        ]

        if request.method == "POST":
            user_id    = int(request.form.get("user_id", 0))
            q_ids      = [int(x) for x in request.form.getlist("question_ids") if x]
            chat_id    = request.form.get("chat_id", "").strip()

            if not user_id or not q_ids or not chat_id:
                flash("Selecciona un usuario y al menos una pregunta.", "error")
                return render_template("send_questions.html",
                                       users=user_list, questions=q_list,
                                       selected_user=user_id, selected_q=[])

            # Create a session with the hand-picked questions
            session = SessionModel(
                user_id      = user_id,
                date         = datetime.utcnow(),
                status       = SessionStatus.PENDING,
                question_ids = q_ids,
                current_question_index = 0,
            )
            db.add(session)
            db.flush()
            engine = SessionEngine(db)

            # Send the FIRST question immediately
            first_q = db.query(Question).get(q_ids[0])
            if not first_q:
                flash("Pregunta no encontrada.", "error")
                return redirect(url_for("dashboard.send_questions"))

            # Build text (reuse the same format as the bot)
            from handlers.conversations import _format_question
            text = _format_question(first_q, 1, len(q_ids))

            ok, err = send_telegram_message(chat_id, text, parse_mode="Markdown")

            if ok:
                # Mark session as in-progress
                engine.start_session(session.id)
                db.commit()

                # Tell the running bot which question is awaiting an answer
                arm_bot_state(chat_id, first_q.id, session.id)

                n = len(q_ids)
                flash(
                    f"✅ {'Pregunta enviada' if n == 1 else f'Sesión de {n} preguntas iniciada'} "
                    f"a {request.form.get('user_name', chat_id)}.",
                    "success",
                )
            else:
                db.rollback()
                invite_link = _invite_link()
                if err and "can't initiate conversation" in (err or "").lower():
                    flash(
                        "⚠️ El rep debe abrir una conversación con el bot primero. "
                        f"Comparte este enlace para que abra el bot: {invite_link}",
                        "error",
                    )
                else:
                    flash(
                        "⚠️ No se pudo enviar el mensaje por Telegram. "
                        + (f"Detalle: {err}" if err else "Verifica TELEGRAM_BOT_TOKEN y chat_id."),
                        "error",
                    )

            return redirect(url_for("dashboard.send_questions"))

        return render_template("send_questions.html",
                               users=user_list, questions=q_list,
                               selected_user=None, selected_q=[])
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# REPORTES — snapshots semanales (WoW), generación automática viernes 10:00 CST (bot)
# ═════════════════════════════════════════════════════════════════════════════


@bp.route("/reportes")
@login_required
def reportes_index():
    db = SessionLocal()
    try:
        rows = (
            db.query(TeamReportSnapshot)
            .order_by(desc(TeamReportSnapshot.period_end_utc))
            .limit(52)
            .all()
        )
        items = []
        for r in rows:
            pl = r.payload or {}
            meta = pl.get("meta") or {}
            items.append(
                {
                    "id": r.id,
                    "period_label": meta.get("period_label_es") or str(r.period_start_utc.date()),
                    "created_at": r.created_at,
                }
            )
        return render_template("reportes.html", reports=items)
    finally:
        db.close()


@bp.route("/reportes/<int:rid>")
@login_required
def reportes_detail(rid: int):
    from config import Config
    from services.report_email import weekly_report_recipients

    db = SessionLocal()
    try:
        r = db.query(TeamReportSnapshot).get(rid)
        if not r:
            flash("Reporte no encontrado.", "error")
            return redirect(url_for("dashboard.reportes_index"))
        import os as _os
        resend_ready = bool((getattr(Config, "RESEND_API_KEY", "") or _os.getenv("RESEND_API_KEY") or "").strip())
        smtp_ready = bool((Config.SMTP_HOST or "").strip() and (Config.SMTP_USER or "").strip() and (Config.SMTP_PASSWORD or "").strip())
        email_ctx = {
            "smtp_ready": smtp_ready or resend_ready,
            "smtp_host": ("Resend (HTTPS)" if resend_ready else (Config.SMTP_HOST or "")),
            "mail_from": Config.WEEKLY_REPORT_EMAIL_FROM or "",
            "recipients": weekly_report_recipients(),
            "enabled": bool(Config.WEEKLY_REPORT_EMAIL_ENABLED),
            "provider": ("resend" if resend_ready else ("smtp" if smtp_ready else "none")),
        }
        return render_template("reportes_detail.html", snap=r, payload=r.payload or {}, email=email_ctx)
    finally:
        db.close()


def _report_url_for(snap_id: int) -> str:
    from config import Config

    base = (getattr(Config, "BASE_URL", "") or "").rstrip("/")
    return f"{base}/reportes/{snap_id}" if base else url_for("dashboard.reportes_detail", rid=snap_id)


@bp.route("/reportes/<int:rid>/email", methods=["POST"])
@login_required
def reportes_email_send(rid: int):
    import copy
    import os as _os2

    from config import Config
    from services.report_email import (
        render_weekly_report_html,
        send_weekly_report_email,
        weekly_report_recipients,
        _plain_text_summary,
    )

    try:
        mode = (request.form.get("mode") or "test").strip().lower()

        # ── 1. Leer snapshot de BD y cerrar sesión antes de cualquier I/O de red ──
        db = SessionLocal()
        try:
            r = db.query(TeamReportSnapshot).get(rid)
            if not r:
                flash("Reporte no encontrado.", "error")
                return redirect(url_for("dashboard.reportes_index"))
            payload = copy.deepcopy(r.payload) if r.payload else {}
            snap_id = r.id
        finally:
            db.close()

        # ── 2. Verificar proveedor de email ──
        _resend_ok = bool((getattr(Config, "RESEND_API_KEY", "") or _os2.getenv("RESEND_API_KEY") or "").strip())
        _smtp_ok = bool(
            (Config.SMTP_HOST or "").strip()
            and (Config.SMTP_USER or "").strip()
            and (Config.SMTP_PASSWORD or "").strip()
        )
        if not _resend_ok and not _smtp_ok:
            flash(
                "Sin proveedor de email configurado. "
                "Define RESEND_API_KEY (recomendado) o SMTP_HOST+SMTP_USER+SMTP_PASSWORD.",
                "error",
            )
            return redirect(url_for("dashboard.reportes_detail", rid=rid))

        # ── 3. Destinatarios y asunto ──
        app_name = os.getenv("APP_NAME", "Sales Coach")
        if mode == "full":
            recipients = weekly_report_recipients()
            subject_prefix = f"{app_name} — Reporte semanal"
        else:
            to_raw = (request.form.get("to") or "").strip()
            if not to_raw:
                flash("Ingresa un correo destino para la prueba.", "error")
                return redirect(url_for("dashboard.reportes_detail", rid=rid))
            recipients = [x.strip() for x in to_raw.split(",") if x.strip()]
            subject_prefix = f"[PRUEBA {app_name}] Reporte semanal"

        label = (payload.get("meta") or {}).get("period_label_es", "Semana")
        subject = f"{subject_prefix} ({label})"

        # ── 4. Renderizar HTML (dentro del try para capturar errores de template) ──
        report_url = _report_url_for(snap_id)
        html = render_weekly_report_html(payload, snap_id, report_url)
        text = _plain_text_summary(payload, report_url)
        mail_from = (Config.WEEKLY_REPORT_EMAIL_FROM or "").strip() or "onboarding@resend.dev"

        # ── 5. Enviar ──
        send_weekly_report_email(
            html_body=html,
            text_body=text,
            subject=subject,
            mail_from=mail_from,
            recipients=recipients,
        )

        if mode == "full":
            flash(f"Reporte enviado a {len(recipients)} destinatarios.", "success")
        else:
            flash(f"Correo de prueba enviado a {', '.join(recipients)}.", "success")

    except Exception as exc:
        log.exception("reportes_email_send error (rid=%s)", rid)
        flash(f"Error al enviar: {exc}", "error")

    return redirect(url_for("dashboard.reportes_detail", rid=rid))


@bp.route("/reportes/generar", methods=["POST"])
@login_required
def reportes_generar():
    """Regenera o crea el snapshot de la última semana cerrada (útil si el bot no corrió)."""
    from services.team_performance_report import (
        last_completed_week_bounds_cst,
        bounds_to_naive_utc,
        build_team_performance_report,
        persist_report,
        get_report_payload_for_period,
    )

    p_start_cst, p_end_cst = last_completed_week_bounds_cst()
    period_start, period_end = bounds_to_naive_utc(p_start_cst, p_end_cst)
    prev_start_cst = p_start_cst - timedelta(days=7)
    prev_end_cst = p_start_cst
    prev_start, prev_end = bounds_to_naive_utc(prev_start_cst, prev_end_cst)

    db = SessionLocal()
    try:
        prev_payload = get_report_payload_for_period(db, prev_start, prev_end)
        payload = build_team_performance_report(
            db, period_start, period_end, prev_start, prev_end, prev_payload
        )
        snap, created = persist_report(db, period_start, period_end, payload)
        if created:
            flash("Reporte semanal generado y guardado.", "success")
        else:
            flash("Ese reporte ya existía; se mantiene el snapshot previo.", "success")
        return redirect(url_for("dashboard.reportes_detail", rid=snap.id))
    except Exception as exc:
        db.rollback()
        flash(f"No se pudo generar el reporte: {exc}", "error")
        return redirect(url_for("dashboard.reportes_index"))
    finally:
        db.close()
